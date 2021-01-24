from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
from selenium import webdriver

import requests

from bs4 import BeautifulSoup as bs

from random import randint

import openpyxl

import time

import datetime

import re

import random

from collections import Counter

import os

import pandas as pd


# delay 주는법 driver.implicitly_wait(3)

# Setting


# keyword : 띄어쓰기는 +로 연결


keyword = "ベビーシート"

###########################################################################

now = datetime.datetime.now()  # 현재 시각 출력 ==> 크롤링하는데 걸린 시간

start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환


URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"

# URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"

URL_SEL = "https://www.amazon.co.jp/"


# 크롬드라이버 위치 설정

driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')


# Session & header 설정

session = requests.Session()

session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
          q =0.9,imgwebp,*/*;q=0.8"}

###########################################################################


# Q excel_str_to_int / str_to_num 함수의 의미

# => 여기서는 쓰이는 함수는 아닌데 의미는 무엇이냐면 파이썬에서 str(1)과 int(1)은 다른 의미를 갖음

# "1" <<의 경우 문자로취급 1 <<의 경우 숫자로 취급함.

# 그래서 문자로된 숫자를 찾아서 엑셀내에서 int형으로 바꿔주는 함수가 str_to_int


# def excel_str_to_int(ws):

#data = []
#
#result_data = []
#
# for col in ws.rows:
#
#        data.append(col[3].value)
#
#    for k in data[1:]:
#
#        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
#
#            result_data.append(int(k))
#
#        else:
#
#            result_data.append(0)
#
#    return result_data


########################

######url 저장(안씀)#####

########################


# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?

# => URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! 근데 이건 안쓸걸요???


# def save_url(data):

#   global SAVE_DIR, J_name
#
#   # wb = load(SAVE_DIR) #load
#
#   wb = openpyxl.Workbook()  # 워크북 생성
#
#   wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
#
#   ws = wb.worksheets[0]
#
#   try:
#
#       for k in range(len(data)):
#
#           ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
#
#           ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
#
#           # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
#
#           ws['C' + str(k+2)] = data[k]
#
#   finally:
#
#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
#
#       wb.close()  # 워크북 닫기


# def str_to_num(word):

#   # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
#
#   p = re.sub("[^0-9]", "", word)
#
#   return p  # Q 이 반환값이 어디서 쓰이나요?
#
#   # => 이것도 그냥 엑셀 처리할때 쓰던건데 여기선 쓰이지 않습니다!_!


#############################################

##########title / abstract(contents)#########

#############################################


# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?

# => 아니욤 이건 예전에 뉴욕타임즈 기사를 수집하던 크롤러라서 제목이 저렇게 되어있는거에요

# 제목이랑 요약된 내용을 수집하겠다는 의미입니다. ㅎㅎ


def find_title_abstract(url_set, year_list):

    global URL_SEL, driver

    # non_abs_url = []  # Q 이건 왜 크롤링?

    # => 뉴욕타임즈에 기사가 없이 사진이 올라와있는 경우가 있어서 따로 url 수집하려고 만들어논 녀석 ==> 안씀

    title_data = []

    abstract_data = []

    authors_data = []

    year_data = []

    # Q year list를 여기서 print 하는 이유? => year가 잘 수집되고있나 중간에 체킹할려고 했던건데 지워도 무관!
    print(year_list)

    for n, journal in enumerate(url_set):

        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴

        driver.get(URL_SEL + journal)

        if 'abs' in journal:

            abstract = driver.find_element_by_class_name('abstractSection')

            if not abstract.text:

                non_abs_url.append(URL_SEL + journal)

            else:

                title = driver.find_element_by_class_name('hlFld-Title')

                authors = driver.find_element_by_class_name('artAuthors')

                title_data.append(title.text)

                abstract_data.append(abstract.text)

                authors_data.append(authors.text)

                # Q year_data는 driver.find로 찾을 필요 없나요?

                # => year는 각각의 기사로 굳이 들어가지않고 페이지 화면에서 수집했기 때문에 따로 수집하지 않음.

                # => 초반에 아예 가지고 있음 페이지 크롤링할때

                year_data.append(year_list[n])

                driver.implicitly_wait(randint(1, 3))

        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘

    print(year_data)

    return title_data, year_data, authors_data, abstract_data  # non_abs_url


#####################

#######url_1#########

#####################


def page_url(source):

    print("시작")

    url_list = []

    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?

    soup = bs(source, 'html.parser')

    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")

    for k in li_list:

        url_list.append(k.attrs['href'])

    try:

        next_butt = soup.find_all('li', class_='a-last')

        if next_butt == []:  # Q == []: ==> 다음 페이지가 아직 더 남았다.

            return url_list, 0

        next_link = next_butt[0].find_all('a')[0].attrs['href']

    except:

        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?) / 다음페이지가 없으니까 끝내라.

        return url_list, 0

    print(len(url_list))

    return url_list, next_link

    #     is_abstract = k.get_text()

    #     if is_abstract == 'Abstract':

    #         url_list.append(k.attrs['href'])

    #     else :

    #         if not 'ref' in k.attrs['href']:

    #             if not k.attrs['href'] in url_list:

    #                 non_abs_url.append(k.attrs['href'])

    # return url_list, non_abs_url


#####################

#######url_2#########

#####################


def main(URL):

    url_list = []

    next_link = URL.format(keyword)

    while (next_link):

        driver.get(next_link)

        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드

        _url, next_link = page_url(source)

        if next_link != 0:

            next_link = URL_SEL + str(next_link)

        url_list += _url

        time.sleep(randint(1, 5))

    return url_list


url_list = main(URL.format(keyword))

print("최종 url")

print(url_list)


###################

######엑셀 저장#####

###################


def save_excel(data):

    global SAVE_DIR, J_name

    # wb = load(SAVE_DIR) #load

    wb = openpyxl.Workbook()

    wb.save('{}.xlsx'.format(J_name))

# wb.save('{}.xlsx'.format(J_name)) ==> 원래 코드
# wb.save('C:/myfolder/....../Amazon_crawling_test_1_0912.xlsx') ==> 경로 지정해준 코드

    ws = wb.worksheets[0]

    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]

    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']

    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정

    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정

        ws[data_header[i]+str(1)] = data_header_name[i]

    try:

        for k in range(len(data)):

            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...

            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨

            ws['B' + str(k+2)] = J_name

            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2

                # data_header[2] (==C) + str3....

                # data_header[3] (==D) + str2

                # data_header[3] (==D) + str3....

                ws[data_header[j + 2]+str(k + 2)] = m

    finally:

        wb.save('{}.xlsx'.format(J_name))

        wb.close()


#####################

######QnA List#######

#####################

# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?
# savve_url ==> URL만 있는 data를 저장할때 쓰기위해 만드것입니다. 인풋 데이터가 url이 들어갈거에요! ==> 안씀

# Q1-2. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?

# 제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용

# (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# ==> 엑셀 파일 다룰때 쓰는 파일 이었는데 클리앙에 이상한 데이터가 너무 많아서 쓰던 함수 여기선 무관! ==> 안씀


####################################################################################################################################

# Q2. page_url함수와 main함수 차이
# url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?

# page_url 함수의 목적
# ==> 처음 검색했을떄 뜨는 항목들의 url을 수집하여 url 리스트와 다음페이지가 존재한다면 다음페이지로 넘어가는 url을 return한다. (_url, next_link)

# 다음페이지가 존재한다면(==next_link가 0이아니라면) 다시한번 page_url함수에 다음 page가 들어가서 항목 url을 긁어온다.

# 아까 C언어만 쓰레기값으로 0을 리턴한다는것 해결 0을 리턴하는경우 => 다음 page가 없을경우 (ex) 아마존에 어떤것을 검색했는데 3페이지까지 뜬다면 4페이지째에는 next_link에 0을)

# main 함수의 목적
# ==> 반환시켜서 next_link가 0일때는 더이상 page_url이 작동하지 않는 구조 이일을 main 함수에서 해주고있음.

####################################################################################################################################

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# ==> 운영체제의 차이

####################################################################################################################################

# Q4. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?

# non_abs_url 리스트의 목적성이 무엇인가요?

# ==> 이것도 뉴욕타임즈에서 쓰던것이라서 있는것임! 뉴욕타임즈에 기사의 요약본은 없고 기사 사진을 찍어놓은것들이 있는데 그것들의 url만 따로 수집하기위해 만들어 놓았음. ==>안씀

####################################################################################################################################

# Q5. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# 음... 무슨 말인지 이해를 못하겠음 ㅠㅠ

####################################################################################################################################

# Q6. next button == []: / 반환값에서 0의 의미?

# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

####################################################################################################################################

# Q7. 맨 처음에 날짜 출력하는 목적이 타임스탬프 때문인건가요?

# 그냥 크롤링 돌리는데 얼마나 걸리는지 궁금해서 만든건데 뒷부분을 삭제한것 같네욤

# now = datetime.datetime.now()  # 현재 시각 출력

# start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# now = datetime.datetime.now()  # 현재 시각 출력

# end_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

# print(end_time-start_time) 하면 아마 걸린시간 나왔던걸로 기억??

####################################################################################################################################

# Q8. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?

# 할수 있습니다. save_url 함수에서

# '''

# finally:

#       wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장

#       wb.close()  # 워크북 닫기

# '''

# 마지막 이부분보면 {}_url.xlsx.format(J_name) 으로 되어있는데 이것을

# 'C:/myfolder/....../제목.xlsx' 이런식으로 수정하면 되겠지용?
