from selenium import webdriver
import requests
from bs4 import BeautifulSoup as bs
from random import randint
import openpyxl
import time
import datetime
import re
import random
from collections import Counter
import os
import pandas as pd

# delay 주는법 driver.implicitly_wait(3)
# Setting

# keyword : 띄어쓰기는 +로 연결

keyword = "ベビーシート"

###########################################################################
###########################################################################
now = datetime.datetime.now()  # 현재 시각 출력
start_time = now.strftime('%Y-%m-%d %H:%M:%S')  # strftime(): 형식 문자열 반환

URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss"
#URL = "https://www.amazon.co.jp/s?k={}&ref=nb_sb_noss_2"
URL_SEL = "https://www.amazon.co.jp/"

# 크롬드라이버 위치 설정 # Q C드라이브 왜 안함?
driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')

# Session & header 설정
session = requests.Session()
session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
         q=0.9,imgwebp,*/*;q=0.8"}
###########################################################################
###########################################################################

# Q excel_str_to_int / str_to_num 함수의 의미


def excel_str_to_int(ws):
    data = []
    result_data = []
    for col in ws.rows:
        data.append(col[3].value)
    for k in data[1:]:
        if str(k).isdigit():  # 문자열이 숫자로 구분되어 있는지 판별
            result_data.append(int(k))
        else:
            result_data.append(0)
    return result_data

###################
######url 저장#####
###################

# Q 코드로 보면 데이터 다 저장하는건데 url만 저장하는게 맞나요?


def save_url(data):
    global SAVE_DIR, J_name
    # wb = load(SAVE_DIR) #load
    wb = openpyxl.Workbook()  # 워크북 생성
    wb.save('{}_url.xlsx'.format(J_name))  # J_name이라는 파일 제목으로 저장
    ws = wb.worksheets[0]
    try:
        for k in range(len(data)):
            ws['A' + str(k + 2)] = k + 1  # A2 = 1, A3 = 2 ==> ID
            ws['B' + str(k+2)] = J_name  # B2 = J Name, B3 = J Name ==> 파일명
            # C2 = data[0], C3 = data[1] ==> [title / year / authors / abstract]
            ws['C' + str(k+2)] = data[k]
    finally:
        wb.save('{}_url.xlsx'.format(J_name))  # 내용 저장 차원에서 다시 저장
        wb.close()  # 워크북 닫기


def str_to_num(word):
    # re.sub ==> 치환 #Q [^0-9] => Q 공백으로 치환하는건 알겠는데 word는 치환할게 없지 않나요?
    p = re.sub("[^0-9]", "", word)
    return p  # Q 이 반환값이 어디서 쓰이나요?


###################################
##########title / abstract#########
###################################

# Q 여기서 abstract의 의미가 이미지같은 확장파일을 의미하나요?


def find_title_abstract(url_set, year_list):
    global URL_SEL, driver
    non_abs_url = []  # Q 이건 왜 크롤링?
    title_data = []
    abstract_data = []
    authors_data = []
    year_data = []
    print(year_list)  # Q year list를 여기서 print 하는 이유?
    for n, journal in enumerate(url_set):
        # Q main1 에서 requests.get과의 차이점 ==> request.get으로는 단순정보만 가져오기때문에 셀레니움 driver.get으로 가져옴
        driver.get(URL_SEL + journal)
        if 'abs' in journal:
            abstract = driver.find_element_by_class_name('abstractSection')
            if not abstract.text:
                non_abs_url.append(URL_SEL + journal)
            else:
                title = driver.find_element_by_class_name('hlFld-Title')
                authors = driver.find_element_by_class_name('artAuthors')
                title_data.append(title.text)
                abstract_data.append(abstract.text)
                authors_data.append(authors.text)
                # Q year_data는 driver.find로 찾을 필요 없나요?
                year_data.append(year_list[n])
                driver.implicitly_wait(randint(1, 3))
        time.sleep(0.1 * random.randint(4, 10))  # 4~10초 사이 랜덤한 시간으로 쉬어줘
    print(year_data)

    # 여기서 non_abs_url 목적
    return title_data, year_data, authors_data, abstract_data, non_abs_url


#####################
#######url_1#########
#####################

def page_url(source):
    print("시작")
    url_list = []
    # Q main1에서는 파싱을 article,title 추출할때 했는데 main2에서는 url란에 있는 이유?
    soup = bs(source, 'html.parser')
    li_list = soup.find_all('a', class_="a-link-normal a-text-normal")
    for k in li_list:
        url_list.append(k.attrs['href'])
    try:
        next_butt = soup.find_all('li', class_='a-last')
        if next_butt == []:  # Q == []:와 반환값에서 0의 의미?
            return url_list, 0
        next_link = next_butt[0].find_all('a')[0].attrs['href']
    except:
        # return 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가?)
        return url_list, 0
    print(len(url_list))
    return url_list, next_link

    #     is_abstract = k.get_text()
    #     if is_abstract == 'Abstract':
    #         url_list.append(k.attrs['href'])
    #     else :
    #         if not 'ref' in k.attrs['href']:
    #             if not k.attrs['href'] in url_list:
    #                 non_abs_url.append(k.attrs['href'])
    # return url_list, non_abs_url

###################
######엑셀 저장#####
###################


def save_excel(data):
    global SAVE_DIR, J_name
    # wb = load(SAVE_DIR) #load
    wb = openpyxl.Workbook()
    wb.save('{}.xlsx'.format(J_name))
    ws = wb.worksheets[0]
    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]
    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']
    ws['A' + str(1)] = data_header_name[0]  # A1 ==> ID에 배정
    for i in range(len(data_header_name)):  # 1행에 입력되는 데이터들 각각 배정
        ws[data_header[i]+str(1)] = data_header_name[i]
    try:
        for k in range(len(data)):
            ws['A' + str(k + 2)] = k + 1  # ID ==> A2 = 1, A3 = 2...
            # Q main1.py에서는 없었던 코든데 넣은 이유 ==> 외부변수로 받아와야 됨
            ws['B' + str(k+2)] = J_name
            for j, m in enumerate(data[k]):  # data_header[2] (==C) + str2
                # data_header[2] (==C) + str3....
                # data_header[3] (==D) + str2
                # data_header[3] (==D) + str3....
                ws[data_header[j + 2]+str(k + 2)] = m
    finally:
        wb.save('{}.xlsx'.format(J_name))
        wb.close()

# url selenium 버전?

#####################
#######url_2#########
#####################


def main(URL):
    url_list = []
    next_link = URL.format(keyword)
    while (next_link):
        driver.get(next_link)
        source = driver.page_source  # 브라우저에 보이는 그대로의 HTML코드
        _url, next_link = page_url(source)
        if next_link != 0:
            next_link = URL_SEL + str(next_link)
        url_list += _url
        time.sleep(randint(1, 5))
    return url_list


url_list = main(URL.format(keyword))
print("최종 url")
print(url_list)

#####################
######QnA List#######
#####################
# Q1. save_url함수와 save_excel 함수 차이 ==> 저장하는 format부분만 다르고 내용은 똑같아보이는데 굳이 나눠서 저장하는 이유가 있나요?

# Q2. page_url함수와 main함수 차이 ==> url크롤링하는 파트로 보이는 부분이 2갠데(url_1과 url_2로 표기 해놨음) 둘의 목적성이 어떻게 다른가요?
# ㄴ제가 생각하기에 전체적인 로직이 url_1은 html긁어와서 url_list에 그 값들을 추가하고 next link를 계속 넘기는 파트고,
#   url_2는 next link가 없을 경우에 url_list를 반환하고 종료하는 파트인 것 같은데 세부적으로 코드 돌아가는 로직이 궁금합니다ㅠ,ㅠ

# Q3. 크롬 드라이버 경로 지정할 때 "C:" 안쓰신 것 같은데 이게 운영체제의 차이인가요? 저는 안쓰면 아예 안되네용

# Q4. save_url파트에서 excel_str_to_int / str_to_num 함수의 존재 목적이 무엇인가요?
# ㄴ제가 생각하기에 엑셀 서식상 오류가 생기니까 [^0-9]같이 특정 특수기호나 문자열들을 숫자로 바꾸려는거같은데 세부적인 코드 설명이 필요해용
#   (그리구 이 코드에서 반환값들은 명목상 있는거고 실제로 쓰이는건 아닌거죠?)

# Q5. title / abstract 파트에서 abstract의 의미가 이미지같은 확장파일을 의미 하나요?
#     non_abs_url 리스트의 목적성이 무엇인가요?

# Q6. main1 파일에서는 파싱을 article,title파트에서 했는데 main2파일에서는 url파트에 있는 이유? (어디서 하든 크게 상관 없나요?)

# Q7. next button == []: / 반환값에서 0의 의미?
# ㄴ return에서의 0 ==> 함수가 에러없이 끝났다는 것을 운영체제에 알려주는 용도(C언어에서만 그런거 아닌가욥?)

# Q8. 맨 처음에 날짜 출력하는 목적이 타임스탬프(?) 때문인건가요?

# Q9. 엑셀 저장 경로는 따로 설정 못하나요?? 설정 안하면 자동으로 파이썬 저장 경로에 저장이 되는건가요?
