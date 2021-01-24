from selenium import webdriver
import requests
from bs4 import BeautifulSoup as bs
from random import randint
import openpyxl
import time
import datetime
import re
import random
from collections import Counter
import os
import pandas as pd

# delay 주는법 driver.implicitly_wait(3)
# Setting

# keyword : 띄어쓰기는 +로 연결

keyword = "fake+sock"

###########################################################################
###########################################################################
# 크롤링 시간 측정
now = datetime.datetime.now() #현재 시간 출력
start_time = now.strftime('%Y-%m-%d %H:%M:%S') #형식 문자열 반환

URL = "https://www.amazon.com/s?k={}&ref=nb_sb_noss_2"
URL_SEL = "https://www.amazon.com/"

# 크롬 드라이버 경로 지정
driver = webdriver.Chrome('C:/Users/Becky/chromedriver_win32/chromedriver.exe')

# Session & header 설정
session = requests.Session()
session.headers = {"User-Agent": "Chrome/68.0 (Macintosh; Intel Win 10 10_9_5)\
         WindowsWebKit 3.80.36 (KHTML, like Gecko) Chrome",
                   "Accept": "text/html,application/xhtml+xml,application/xml;\
         q=0.9,imgwebp,*/*;q=0.8"}
###########################################################################
###########################################################################

#####url 수집#####
def main(URL):
    url_list = []
    next_link = URL.format(keyword)
    try:
        while (next_link):
            driver.get(next_link)
            source = driver.page_source
            _url, next_link = page_url(source)
            next_link = URL_SEL + next_link
            url_list += _url
            time.sleep(randint(1, 5))
        return url_list
    except:
        return url_list

#####url 수집#####
def components(source):
    print("url 크롤링 시작")
    url_list = []
    soup = bs(source, 'html.parser')
    u_list = soup.find_all(
        'div', class_="a-section a-spacing-none a-spacing-top-small")
    for url in u_list:
        urll = url.find_all('a', class_='a-link-normal a-text-normal')
        for u in urll:
            url_list.append(u.attrs['href'])

##### image / brand / title / review rating / original price / price reduction 수집 #####
    print("components 크롤링 시작")
    image_list = []
    brand_list = []
    title_list = []
    rating_list = []
    originalprice_list = []
    pricereduction_list = []
    soup = bs(source, 'html.parser')

    # image (이미지 필요)
    i_list = soup.find_all(
        'div', class_="a-section a-spacing-none s-image-overlay-black")
    for img in i_list:
        image = img.find_all('div', class_='a-section aok-relative s-image-tall-aspect')
        for i in image:
            image_list.append(i.attrs['img src'])

    # brand (링크 필요x)
    b_list = soup.find_all(
        'div', class_="a-row a-size-base a-color-secondary")
    for brd in b_list:
        brand = brd.find.all('span', class_='a-size-base-plus a-color-base')
        for b in brand:
        brand_list.append(b.attrs['span'])

    # title (링크 필요o)
    t_list = soup.find_all(
        'h2', class_="a-size-mini a-spacing-none a-color-base s-line-clamp-2")
    for ttl in t_list:
        title = ttl.find_all('a', class_='a-link-normal a-text-normal')
        for t in title:
            title_list.append(t.attrs['href'])

    # rating (링크 필요o)
    r_list = soup.find_all(
        'div', class_="a-section a-spacing-none a-spacing-top-micro")
    for rat in r_list:
        rating = rat.find_all('a', class_="a-row a-size-small")
        for r in rating:
            rating_list.append(r.attrs['div'])

    # original price (링크 필요x)
    op_list = soup.find_all(
        'div', class_="a-row")
    for opc in op_list:
        original = opc.find_all('span', class_="a-price")
        for o in original:
            originalprice_list.append(o.attrs['span'])

    # price reduction (링크 필요x)
    pr_list = soup.find_all(
         'div', class_="a-row")
    for prd in pr_list:
        reduction = prd.find.all('span', class_="a-price a-text-price")
        for p in reduction:
            pricereduction_list.append(p.attrs['span'])

    try:
        next_butt = soup.find_all('li', class_='a-last')
        next_link = next_butt[0].find_all('a')[0].attrs['href']
    except:
        return url_list, image_list, title_list, rating_list, originalprice_list, pricereduction_list, next_link, 0 #다음페이지가 있으면(=0이 아니라면) 다시한번 다음 page 가서 url 긁어와라

    print(len(url_list),len(image_list), len(title_list), len(rating_list), len(originalprice_list), len(pricereduction_list))

    return url_list, image_list, title_list, rating_list, originalprice_list, pricereduction_list, next_link

    #     is_abstract = k.get_text()
    #     if is_abstract == 'Abstract':
    #         url_list.append(k.attrs['href'])
    #     else :
    #         if not 'ref' in k.attrs['href']:
    #             if not k.attrs['href'] in url_list:
    #                 non_abs_url.append(k.attrs['href'])
    # return url_list, non_abs_url


def save_excel(data):
    global SAVE_DIR, J_name
    # wb = load(SAVE_DIR) #load
    wb = openpyxl.Workbook()
    wb.save('{}.xlsx'.format(J_name))
    ws = wb.worksheets[0]
    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]
    data_header_name = ['id_num', 'J_name',
                        'title', 'year', 'authors', 'abstract']
    ws['A' + str(1)] = data_header_name[0]
    for i in range(len(data_header_name)):
        ws[data_header[i]+str(1)] = data_header_name[i]
    try:
        for k in range(len(data)):
            ws['A' + str(k + 2)] = k + 1
            ws['B' + str(k+2)] = J_name
            for j, m in enumerate(data[k]):
                ws[data_header[j + 2]+str(k + 2)] = m
    finally:
        wb.save('{}.xlsx'.format(J_name))
        wb.close()

######엑셀 저장#######
url_list = main(URL.format(keyword))

# Create a Pandas dataframe from the data.
df = pd.DataFrame(
    {
        "URL" : url_list
    }
)

# Create a Pandas Excel writer using XlsxWriter as the engine.
writer = pd.ExcelWriter('url_save_test1.xlsx', engine='xlsxwriter', options={'strings_to_urls': False})

# Convert the dataframe to an XlsxWriter Excel object.
df.to_excel(writer)
writer.save()
print("===============종료============")
print(url_list)
