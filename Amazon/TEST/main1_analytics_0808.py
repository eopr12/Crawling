import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import pandas as pd
import json
import re
import sys
import time
import random

# 0 ~ 28
# https://www.clien.net/service/search?q=산업%20수학&sort=recency&p=0&boardCd=&isBoard=false

URL_SEL = "https://www.clien.net/"

###################
######엑셀 저장#####
###################

# Q 파일 상단 xlsxwriter VS. openpyxl 차이 두개 다 쓰는 이유?
# ==> openpyxl은 불러오기 쓰기 모두 가능 / xlsxwriter는 쓰기만 가능
# ==> openpyxl은 cell(1,1)부터 시작(range 1부터 n+1까지) / xlsxwriter는 cell(0,0)부터 시작(range 그냥 n)
<<<<<<< HEAD
# ==> xlsxwriter가 처리 속도 더 빠름
# ==> xlsxwriter은 한국어 reading 오류가 자주 남
=======
# xlsxwriter가 처리 속도 더 빠름
# xlsxwriter는 한글이 이상하게 처리되는 경향이 있음 ==> 영어에 최적화
>>>>>>> b68e3936d30249d5ccf80cd2e2e25d728dc34a25


def save_excel(data, file_name):
    # wb = load(SAVE_DIR) #load
    wb = openpyxl.Workbook()  # 워크북 생성
    wb.save('{}.xlsx'.format(file_name))  # file_name이라는 제목으로 엑셀 저장
    # Q 워크시트 생성 wb.active 생략? ==> 엑셀은 default로 sheet가 하나씩 있기 때문에 굳이 활성화를 할 필요는 없음
    ws = wb.worksheets[0]
    # ord ==> 문자의 ASCII 코드값 구하기 / # chr ==> 해당 ASCII 코드에 해당하는 값을 문자로 출력하기
    # Q 왜 하필 콕 집어서 A랑 Z? ==> 엑셀이 데이터 저장하는 방식 A to Z 차례대로 저장하기 때문
    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]  # 차례대로 데이터 저장
    data_header_name = ['ID', 'Title', 'article', 'comment']  # 엑셀 헤더부분 이름 지정
    # A1데이터가 ID에 배정 # Q 바로 밑에 똑같이 대입되는 공식 있는데 굳이 존재하는 이유?
    ws['A' + str(1)] = data_header_name[0]
    for i in range(len(data_header_name)):  # len(List) ==> 리스트 개수
        # 1행에 입력되는 데이터들 입력 ==> A1 = 'ID', B1 = 'Title'....
        ws[data_header[i]+str(1)] = data_header_name[i]
    try:
<<<<<<< HEAD
        for k in range(len(data)):  # data ==> 외부에서 받아오는 데이터
            ws['A' + str(k+2)] = k+1  # 고유식별코드 입력
            # j는 옆에 표시되는 상수 / m은 변수 ==> [0, data[0]] 이런식으로 표기 됨
=======
        for k in range(len(data)):  # Q data ==> 외부에서 받아오는 데이터
            ws['A' + str(k+2)] = k+1  # 고유식별코드 입력 ==> A2 = 1, A3 = 2...
            # j는 옆에 표시되는 상수 / m은 변수 ==> [0, data[k]] 이런식으로 표기 됨
            # data ==> [ A2, B2, C2, D2 ]..............
            # j먼저 들어가서 차례대로 쭉 ==> data_header[1] (==B) + str2 = data[k]
            #   data_header[1] (==B) + str3 = data[k]......
            #   data header[2] (==C) + str2 = data[k]
>>>>>>> b68e3936d30249d5ccf80cd2e2e25d728dc34a25
            for j, m in enumerate(data[k]):
                ws[data_header[j+1]+str(k + 2)] = m
    finally:
        # Q 처음이랑 끝에 save 두번 하는 이유 ==> 처음엔 파일이 없을 수도 없으니까 / 나중엔 저장 차원에서
        wb.save('{}.xlsx'.format(file_name))
        wb.close()  # 워크북 닫기

####1####
###################################
#########article / comment#########
###################################


def find_title_abstract(url_set):
    global URL_SEL, driver  # 셀레니움 드라이버
    article_data = []  # 리스트 처리
    comment_data = []

    # enumerate() ==> for in과 다르게 몇 번째 반복문인지 옆에 나타내고 싶을 때 사용
    for aa, journal in enumerate(url_set):

        rq = requests.get(URL_SEL + journal)  # url속 html 가져오기
        soup = bs(rq.text, 'html.parser')  # html 가져와서 parsing
        tmp_article = soup.find_all(
            'div', class_="post_article")  # 원하는 값 찾기(클래스로 찾기)
        article = ""
        if tmp_article[0].get_text():
            article += tmp_article[0].get_text().strip().replace(  # strip() 문자열 양끝 공백제거 \n 제거
                "\xa0", " ").replace("#CLiOS", "").replace("\r", "")  # Q \xa0랑 CLi0S 왜 공백처리 해야됨? ==> 그냥 뽑았더니 엑셀에 이런 문자가 껴있어서
        else:
            for art in tmp_article:
                tt_art = art.find_all('p')  # p라는 글자를 포함하는 태그를 가져옴
                for tt in tt_art:
                    article += tt.get_text()
        article_data.append(article)  # article list에 추가

        comment = ""  # Q string으로 만든거?
        try:
            tmp_comment = soup.find_all('div', class_="comment_view")
            for com in tmp_comment:
                comment += com.get_text().strip().replace("\xa0", " ").replace("#CLiOS",
                                                                               "").replace("\r", "").replace("%20", " ")  # Q \xa0랑 CLi0S랑 %20 왜 공백처리 해야됨?
                comment += "\n"
            comment_data.append(comment)
        except:
            comment_data.append(comment)
    return article_data, comment_data

#####################
#####url&title#######
#####################


def page_url(soup):
    url_list = []
    title_list = []
    # a라는 글자를 포함하는 태그를 가져옴(클래스로 찾기)
    li_list = soup.find_all('a', class_="subject_fixed")

    for k in li_list:  # Q
        kk = k.attrs['href']
        kkk = k.get_text()
        url_list.append(kk)
        title_list.append(kkk)
    return url_list, title_list

#################
#####Running#####
#################


Data = [[], [], [], []]
ID = []
for num in range(0, 30):
    rq = requests.get(
        "https://www.clien.net/service/search?q=산업%20수학&sort=recency&p={}&boardCd=&isBoard=false".format(num))
    code = rq.status_code
    soup = bs(rq.text, 'html.parser')
    print('===============================')
    print('this page status code : {}'.format(code))  # 상태코드를 출력
    if code == 200:  # ok 응답코드
        print("{}page 크롤링 시작".format(num))
        page_list, title = page_url(soup)
        article, comment = find_title_abstract(page_list)
    if not title in Data[0]:
        Data[0] += title
        Data[1] += article
        Data[2] += comment
        Data[3] += page_list

for i in range(len(Data[0])):
    ID.append(i+1)  # 0,1,2,3,4,........

#######################
#### 엑셀 파일 생성#####
#######################

# Q 파일 상단 xlsxwriter VS. openpyxl 차이 두개 다 쓰는 이유?

df = pd.DataFrame(  # 행과 열이 있는 2차원 tabular data / 열(column)을 dict의 Key로, 행(row)을 dict의 Value로 해서 Data Frame 생성
    {'ID': ID,
     'Title': Data[0],
     'article': Data[1],
     'comment': Data[2],
     'URL': Data[3]
     }
)

writer = pd.ExcelWriter('out.xlsx', engine='xlsxwriter',  # XlsxWriter 엔진으로 Pandas writer 객체 만들기
                        options={'strings_to_urls': False})
# 생성된 Data Frame을 xlsx에 쓰기 / sheet name 지정 가능 ex)df.to_excel(writer, sheet_name='Sheet1')
df.to_excel(writer)
writer.save()  # Pandas writer 객체 저장
