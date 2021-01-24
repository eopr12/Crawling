import requests
from bs4 import BeautifulSoup as bs
import openpyxl
import pandas as pd
import json
import re
import sys
import time
import random

# 0 ~ 28
# https://www.clien.net/service/search?q=산업%20수학&sort=recency&p=0&boardCd=&isBoard=false

URL_SEL = "https://www.clien.net/"

###################
######엑셀 저장#####
###################

# Q 파일 상단 xlsxwriter VS. openpyxl 차이 두개 다 쓰는 이유?


def save_excel(data, file_name):
    # wb = load(SAVE_DIR) #load
    wb = openpyxl.Workbook()  # 워크북 생성
    wb.save('{}.xlsx'.format(file_name))  # file_name이라는 제목으로 엑셀 저장
    ws = wb.worksheets[0]  # Q 워크시트 생성 wb.active 생략?
    # ord ==> 문자의 코드값 구하기 / # chr ==> 코드에 해당하는 값을 문자로 출력하기
    # Q 왜 하필 콕 집어서 A랑 Z?
    data_header = [chr(x) for x in range(ord('A'), ord('Z')+1)]
    data_header_name = ['ID', 'Title', 'article', 'comment']  # 엑셀 헤더부분 서식 지정
    # Q str(1)이 'A'의 문자열 출력? 이하 전반적으로 이해 안감
    ws['A' + str(1)] = data_header_name[0]
    for i in range(len(data_header_name)):
        ws[data_header[i]+str(1)] = data_header_name[i]
    try:
        for k in range(len(data)):
            ws['A' + str(k+2)] = k+1
            for j, m in enumerate(data[k]):
                ws[data_header[j+1]+str(k + 2)] = m
    finally:
        wb.save('{}.xlsx'.format(file_name))
        wb.close()  # 워크북 닫기

###################################
#########article / comment#########
###################################


def find_title_abstract(url_set):
    global URL_SEL, driver  # 셀레니움 드라이버
    article_data = []  # 리스트 처리
    comment_data = []

    # enumerate() ==> for in과 다르게 몇 번째 반복문인지 옆에 나타내고 싶을 때 사용
    for aa, journal in enumerate(url_set):  # Q aa 뭐임?

        rq = requests.get(URL_SEL + journal)  # url속 html 가져오기
        soup = bs(rq.text, 'html.parser')  # html 가져와서 parsing
        tmp_article = soup.find_all(
            'div', class_="post_article fr-view")  # 원하는 값 찾기(클래스로 찾기)
        article = ""
        if tmp_article[0].get_text():
            article += tmp_article[0].get_text().strip().replace(  # strip() 문자열 양끝 공백제거 \n 제거
                "\xa0", " ").replace("#CLiOS", "").replace("\r", "")  # Q \xa0랑 CLi0S 왜 공백처리 해야됨?
        else:
            for art in tmp_article:
                tt_art = art.find_all('p')  # p라는 글자를 포함하는 태그를 가져옴
                for tt in tt_art:
                    article += tt.get_text()
        article_data.append(article)  # article list에 추가

        comment = ""
        try:
            tmp_comment = soup.find_all('div', class_="comment_view")
            for com in tmp_comment:
                comment += com.get_text().strip().replace("\xa0", " ").replace("#CLiOS",
                                                                               "").replace("\r", "").replace("%20", " ")  # Q \xa0랑 CLi0S랑 %20 왜 공백처리 해야됨?
                comment += "\n"
            comment_data.append(comment)
        except:
            comment_data.append(comment)
    return article_data, comment_data

##################
######title#######
##################


def page_url(soup):
    url_list = []
    title_list = []
    # a라는 글자를 포함하는 태그를 가져옴(클래스로 찾기)
    li_list = soup.find_all('a', class_="subject_fixed")

    for k in li_list:
        kk = k.attrs['href']
        kkk = k.get_text()
        url_list.append(kk)
        title_list.append(kkk)
    return url_list, title_list

############
#####ID#####
############


Data = [[], [], [], []]
ID = []
for num in range(0, 30):
    rq = requests.get(
        "https://www.clien.net/service/search?q=산업%20수학&sort=recency&p={}&boardCd=&isBoard=false".format(num))
    code = rq.status_code
    soup = bs(rq.text, 'html.parser')
    print('===============================')
    print('this page status code : {}'.format(code))  # 상태코드를 출력
    if code == 200:  # ok 응답코드
        print("{}page 크롤링 시작".format(num))
        page_list, title = page_url(soup)
        article, comment = find_title_abstract(page_list)
    if not title in Data[0]:
        Data[0] += title
        Data[1] += article
        Data[2] += comment
        Data[3] += page_list

for i in range(len(Data[0])):
    ID.append(i+1)  # ID 산정방식이 이거임?

#######################
#### 엑셀 파일 생성#####
#######################

# Q 파일 상단 xlsxwriter VS. openpyxl 차이 두개 다 쓰는 이유?

df = pd.DataFrame(  # 행과 열이 있는 2차원 tabular data / 열(column)을 dict의 Key로, 행(row)을 dict의 Value로 해서 Data Frame 생성
    {'ID': ID,
     'Title': Data[0],
     'article': Data[1],
     'comment': Data[2],
     'URL': Data[3]
     }
)

writer = pd.ExcelWriter('out.xlsx', engine='xlsxwriter',  # XlsxWriter 엔진으로 Pandas writer 객체 만들기
                        options={'strings_to_urls': False})
# 생성된 Data Frame을 xlsx에 쓰기 / sheet name 지정 가능 ex)df.to_excel(writer, sheet_name='Sheet1')
df.to_excel(writer)
writer.save()  # Pandas writer 객체 저장
